import base64
from flask import Flask, render_template, request
import openpyxl
from waitress import serve
from pandas import pandas as pd
from plotly import graph_objects as go
import matplotlib.pyplot as plt
import numpy as np
from io import BytesIO

app = Flask(__name__)

@app.route('/')
@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/graph', methods=['POST'])
def display_graph():
    #input
    firstName = request.form['firstname']
    lastName = request.form['lastname']
    dateOfBirth = request.form['dateofbirth']

    dataFile = request.files['datafile']
    if dataFile and dataFile.filename.endswith('.xlsx'):
        excel_data = pd.read_excel(dataFile)
        dataGraph = render_graph(excel_data)

    #save user history
    file_path = 'Data/login.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Sheet1']
    next_row = sheet.max_row + 1

    data_to_append = [firstName, lastName, dateOfBirth]

    for col_num, value in enumerate(data_to_append, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)

    workbook.save(file_path)

    return render_template(
        "graph.html",
        title= 'Income and Expenditure',
        customer = f"{firstName} {lastName}",
        graph_html=f"<img src='data:image/png;base64,{dataGraph}'/>"
    )

def render_graph(excel_data):
    _month = []
    _income = []
    _expense = []
    for index, row in excel_data.iterrows():
        _month.append(row['Month'])
        _income.append(row['Income'])
        _expense.append(row['Expenses'])

    data = {
        'Income': _income,
        'Expense': _expense
    }

    x = np.arange(len(_month))  # the label locations
    width = 0.25  # the width of the bars
    multiplier = 0

    fig, ax = plt.subplots(layout='constrained', figsize=(16,9))

    for attribute, measurement in data.items():
        offset = width * multiplier
        rects = ax.bar(x + offset, measurement, width, label=attribute)
        ax.bar_label(rects, padding=3)
        multiplier += 1

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_ylabel('ZAR')
    ax.set_title('Income and Expenditure')
    ax.set_xticks(x + width, _month)
    ax.legend(loc='upper left', ncols=3)

    # Save it to a temporary buffer
    buf = BytesIO()
    fig.savefig(buf, format="png")
    # Embed the result
    data = base64.b64encode(buf.getbuffer()).decode("ascii")
    return data


if __name__ == "__main__":
    # app.run(host="0.0.0.0", port=8000)
    serve(app, host="0.0.0.0", port=4200)